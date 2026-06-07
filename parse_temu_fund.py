#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Temu 账务明细解析器（支持全托管/半托管）
"""

import argparse
import sys
import warnings
from pathlib import Path

import pandas as pd
from openpyxl.comments import Comment

# 忽略 openpyxl 的默认样式警告
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# 抑制 pandas fillna downcast 警告
pd.set_option("future.no_silent_downcasting", True)

REGIONS = ["欧区", "美区", "全球"]


def find_region_file(folder: Path, region: str) -> Path:
    """在目录中查找匹配区域的账单文件，支持关键词模糊匹配。"""
    keywords = {
        "欧区": ["欧区", "欧洲"],
        "美区": ["美区", "美国"],
        "全球": ["全球"],
    }
    # 获取目录下所有 xlsx 文件
    files = list(folder.glob("*.xlsx"))
    for kw in keywords.get(region, [region]):
        for f in files:
            if kw in f.name:
                return f
    return folder / f"{region}账务明细.xlsx"


def _col_width(val, number_format=None):
    if val is None:
        return 0
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        # 按显示格式计算宽度，避免 Python float 全精度导致过宽
        if number_format and "0.00" in number_format:
            s = f"{val:,.2f}"
        elif isinstance(val, int):
            s = str(val)
        else:
            s = f"{val:.4f}"
    else:
        s = str(val)
    return sum(2 if ord(ch) > 127 else 1 for ch in s)


def fmt_worksheet(ws, numeric_cols, df):
    for c_idx, c_name in enumerate(df.columns, start=1):
        if c_name not in numeric_cols:
            continue
        for r_idx in range(2, len(df) + 2):
            cell = ws.cell(row=r_idx, column=c_idx)
            v = str(cell.value).replace(".", "", 1).replace("-", "", 1)
            if isinstance(cell.value, (int, float)) or v.isdigit():
                cell.number_format = "0.00"

    for c_idx in range(1, ws.max_column + 1):
        max_w = 0
        # 只采样表头 + 前 50 行数据，避免异常值把列宽撑得过大
        check_rows = min(ws.max_row, 51)
        for r_idx in range(1, check_rows + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            w = _col_width(cell.value, cell.number_format)
            if w > max_w:
                max_w = w
        col_letter = ws.cell(row=1, column=c_idx).column_letter
        ws.column_dimensions[col_letter].width = min(max(max_w + 2, 8), 30)


def read_sheets(path: Path):
    xl = pd.ExcelFile(str(path))
    names = set(xl.sheet_names)
    is_empty = not ({"交易结算", "结算"} & names)
    is_half = "结算" in names and "交易结算" not in names
    return xl, is_empty, is_half


def read_settlement(xl, is_half: bool):
    if is_half:
        df = pd.read_excel(xl, sheet_name="结算", header=[0, 1])
        cols = []
        for c in df.columns:
            col = (
                c[1] if not (pd.isna(c[1]) or str(c[1]).startswith("Unnamed")) else c[0]
            )
            cols.append(col)
        df.columns = cols
        df = df.rename(
            columns={"结算金额": "金额", "SKU名称": "货品名称", "件数": "数量"}
        )
        df["SKU属性"] = ""
        for col in ["SKU ID", "SKU货号", "货品名称"]:
            df[col] = df[col].astype(str).str.strip().replace("nan", "")
        df["金额"] = pd.to_numeric(df["金额"], errors="coerce").fillna(0)
        df["数量"] = pd.to_numeric(df.get("数量", 0), errors="coerce").fillna(0)

        ship_mask = df["交易类型"].isin(["运费回款", "运费冲回"])
        extra_ship = float(df.loc[ship_mask, "金额"].sum())

        no_sku = (df["SKU ID"] == "") & (~ship_mask)
        extra_refund = float(
            df.loc[no_sku & (df["交易类型"] == "销售冲回"), "金额"].sum()
        )

        df = df[(~ship_mask) & (df["SKU ID"] != "")].copy()
        return df, {"运费回款": extra_ship, "销售冲回_no_sku": extra_refund}

    df = pd.read_excel(xl, sheet_name="交易结算")
    df["金额"] = pd.to_numeric(df["金额"], errors="coerce").fillna(0)
    if "数量" in df.columns:
        df["数量"] = pd.to_numeric(df["数量"], errors="coerce").fillna(0)
    for col in ["SKU ID", "SKU货号", "货品名称", "SKU属性"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()
    return df, {}


def read_compensation(xl, is_half: bool):
    comp_dfs = []
    sku_sheets = {}

    if is_half:
        sheets = [s for s in xl.sheet_names if "支出-履约违规" in s]
        for s in sheets:
            df = pd.read_excel(xl, sheet_name=s)
            if "支出金额" not in df.columns:
                continue
            df["赔付金额"] = pd.to_numeric(df["支出金额"], errors="coerce").fillna(0)
            comp_dfs.append(df[["赔付金额"]])
        extra_comp = (
            sum(float(pd.concat(comp_dfs)["赔付金额"].sum()) for _ in [0] if comp_dfs)
            if comp_dfs
            else 0.0
        )
        return comp_dfs, sku_sheets, extra_comp

    sheets = [
        s
        for s in xl.sheet_names
        if any(k in s for k in ["消费者履约保障", "消费者及履约保障"])
    ]
    for s in sheets:
        df = pd.read_excel(xl, sheet_name=s)
        if "赔付金额" not in df.columns:
            continue
        df["赔付金额"] = pd.to_numeric(df["赔付金额"], errors="coerce").fillna(0)
        for col in ["SKU ID", "SKU货号"]:
            df[col] = df[col].astype(str).str.strip()
        comp_dfs.append(df[["SKU ID", "SKU货号", "赔付金额"]])
        sku_sheets[s] = (
            df.groupby(["SKU ID", "SKU货号"])["赔付金额"].sum().reset_index()
        )
        info = df[
            [c for c in ["SKU ID", "SKU货号", "货品名称", "SKU属性"] if c in df.columns]
        ]
        if not info.empty:
            sku_sheets[f"{s}_info"] = info.drop_duplicates(subset=["SKU ID", "SKU货号"])
    return comp_dfs, sku_sheets, 0.0


def build_pivot(df, comp_dfs, sku_sheets):
    pivot = (
        df.groupby(["SKU ID", "SKU货号", "货品名称", "SKU属性", "交易类型"])["金额"]
        .sum()
        .unstack(fill_value=0)
        .reset_index()
    )

    sheet_map = {}
    for s, d in sku_sheets.items():
        if s.endswith("_info"):
            continue
        short = (
            "售后补寄赔付金额"
            if "售后补寄" in s
            else "售后问题赔付金额"
            if "售后问题" in s
            else f"赔付金额-{s}"
        )
        sheet_map[s] = short
        pivot = pivot.merge(
            d.rename(columns={"赔付金额": short}), on=["SKU ID", "SKU货号"], how="outer"
        )
        pivot[short] = pivot[short].fillna(0)

    # 后续 merge（how="outer"）可能引入新 SKU，导致之前已填充的列重新出现 NaN
    for short in list(sheet_map.values()):
        pivot[short] = pivot[short].fillna(0)

    comp_total = pd.DataFrame(columns=["SKU ID", "SKU货号", "总赔付金额"])
    if comp_dfs:
        c = pd.concat(comp_dfs, ignore_index=True)
        if {"SKU ID", "SKU货号"}.issubset(c.columns):
            comp_total = (
                c.groupby(["SKU ID", "SKU货号"])["赔付金额"]
                .sum()
                .reset_index()
                .rename(columns={"赔付金额": "总赔付金额"})
            )
    pivot = pivot.merge(comp_total, on=["SKU ID", "SKU货号"], how="outer")
    pivot["总赔付金额"] = pivot["总赔付金额"].fillna(0)

    infos = [v for k, v in sku_sheets.items() if k.endswith("_info")]
    if infos:
        info = pd.concat(infos, ignore_index=True).drop_duplicates(
            subset=["SKU ID", "SKU货号"]
        )
        pivot = pivot.merge(
            info, on=["SKU ID", "SKU货号"], how="left", suffixes=("", "_i")
        )
        for col in ["货品名称", "SKU属性"]:
            if f"{col}_i" in pivot.columns:
                pivot[col] = pivot[col].fillna(pivot[f"{col}_i"])
                pivot = pivot.drop(columns=[f"{col}_i"])

    static = ["SKU ID", "SKU货号", "货品名称", "SKU属性"]
    known = list(sheet_map.values()) + ["总赔付金额", "净回款总额"]
    tx = [c for c in pivot.columns if c not in static and c not in known]
    for c in tx:
        pivot[c] = pivot[c].fillna(0)
    pivot["货品名称"] = pivot["货品名称"].fillna("")
    pivot["SKU属性"] = pivot["SKU属性"].fillna("")
    pivot["净回款总额"] = pivot[tx].sum(axis=1) - pivot["总赔付金额"]

    for c in list(sheet_map.values()) + ["总赔付金额"]:
        if c in pivot.columns:
            pivot[c] = -pivot[c]

    back = [
        c
        for c in ["售后补寄赔付金额", "售后问题赔付金额", "总赔付金额", "净回款总额"]
        if c in pivot.columns
    ]
    mid = [c for c in pivot.columns if c not in static + back]
    if "销售回款" in mid and "销售冲回" in mid:
        mid.remove("销售回款")
        mid.remove("销售冲回")
        mid = ["销售回款", "销售冲回"] + mid
    pivot = pivot[static + mid + back]
    return pivot


def build_metrics(df, comp_dfs):
    rows = []
    seen = set()
    for (sku_id, sku_no, name, attr), g in df.groupby(
        ["SKU ID", "SKU货号", "货品名称", "SKU属性"]
    ):
        seen.add((sku_id, sku_no))
        back = g["交易类型"] == "销售回款"
        refund = g["交易类型"] == "销售冲回"
        n_back = int(back.sum())
        n_refund = int(refund.sum())
        sales_val = float(g.loc[back, "金额"].sum())
        sales_only = sales_val
        refund_val = abs(float(g.loc[refund, "金额"].sum()))
        qty = float(g.loc[back, "数量"].sum()) if "数量" in g.columns else 0.0

        comp = pd.concat(comp_dfs, ignore_index=True) if comp_dfs else None
        has_sku = comp is not None and {"SKU ID", "SKU货号"}.issubset(comp.columns)
        sub = (
            comp[(comp["SKU ID"] == sku_id) & (comp["SKU货号"] == sku_no)]
            if has_sku
            else None
        )
        comp_val = (
            float(sub["赔付金额"].sum()) if sub is not None and not sub.empty else 0.0
        )
        comp_cnt = len(sub) if sub is not None else 0

        rows.append(
            {
                "SKU ID": sku_id,
                "SKU货号": sku_no,
                "货品名称": name,
                "SKU属性": attr,
                "销售数量": int(qty),
                "销售回款总额": round(sales_only, 2),
                "回款订单数量": n_back,
                "退款订单数量": n_refund,
                "退货率": n_refund / n_back if n_back else 0.0,
                "赔付订单数量": comp_cnt,
                "赔付率": comp_cnt / n_back if n_back else 0.0,
                "退货总额": round(refund_val, 2),
                "退货金额比例": refund_val / sales_only if sales_only else 0.0,
                "赔付金额": round(comp_val, 2),
                "赔付金额比例": comp_val / sales_val if sales_val else 0.0,
            }
        )

    comp = pd.concat(comp_dfs, ignore_index=True) if comp_dfs else None
    if comp is not None and {"SKU ID", "SKU货号"}.issubset(comp.columns):
        miss = comp[~comp.set_index(["SKU ID", "SKU货号"]).index.isin(seen)][
            ["SKU ID", "SKU货号"]
        ].drop_duplicates()
        for _, r in miss.iterrows():
            sub = comp[
                (comp["SKU ID"] == r["SKU ID"]) & (comp["SKU货号"] == r["SKU货号"])
            ]
            rows.append(
                {
                    "SKU ID": r["SKU ID"],
                    "SKU货号": r["SKU货号"],
                    "货品名称": str(sub["货品名称"].iloc[0])
                    if "货品名称" in sub.columns and len(sub)
                    else "",
                    "SKU属性": str(sub["SKU属性"].iloc[0])
                    if "SKU属性" in sub.columns and len(sub)
                    else "",
                    "销售数量": 0,
                    "销售回款总额": 0.0,
                    "回款订单数量": 0,
                    "退款订单数量": 0,
                    "退货率": 0.0,
                    "赔付订单数量": len(sub),
                    "赔付率": 0.0,
                    "退货总额": 0.0,
                    "退货金额比例": 0.0,
                    "赔付金额": round(float(sub["赔付金额"].sum()), 2),
                    "赔付金额比例": 0.0,
                }
            )
    return pd.DataFrame(rows)


def process_region(path: Path):
    xl, is_empty, is_half = read_sheets(path)
    if is_empty:
        empty_pivot = pd.DataFrame(
            columns=["SKU ID", "SKU货号", "货品名称", "SKU属性", "净回款总额"]
        )
        empty_metrics = pd.DataFrame(
            columns=[
                "SKU ID",
                "SKU货号",
                "货品名称",
                "SKU属性",
                "销售数量",
                "销售回款总额",
                "回款订单数量",
                "退款订单数量",
                "退货率",
                "赔付订单数量",
                "赔付率",
                "退货总额",
                "退货金额比例",
                "赔付金额",
                "赔付金额比例",
            ]
        )
        return empty_pivot, empty_metrics, {}, pd.DataFrame()

    settle, extra = read_settlement(xl, is_half)
    comp_dfs, sku_sheets, extra_comp = read_compensation(xl, is_half)
    extra["赔付总额"] = extra_comp
    pivot = build_pivot(settle, comp_dfs, sku_sheets)
    metrics = build_metrics(settle, comp_dfs)
    return pivot, metrics, extra, settle


def merge_shop(pivot_list, metrics_list):
    all_pivot = pd.concat(pivot_list, ignore_index=True)
    keys = ["SKU ID", "SKU货号", "货品名称", "SKU属性"]
    mp = all_pivot.groupby(keys, as_index=False).sum(numeric_only=False)
    back = [
        c
        for c in ["售后补寄赔付金额", "售后问题赔付金额", "总赔付金额", "净回款总额"]
        if c in mp.columns
    ]
    mid = [c for c in mp.columns if c not in keys + back]
    if "销售回款" in mid and "销售冲回" in mid:
        mid.remove("销售回款")
        mid.remove("销售冲回")
        mid = ["销售回款", "销售冲回"] + mid
    mp = mp[keys + mid + back]

    non_empty = [m for m in metrics_list if not m.empty]
    am = pd.concat(non_empty, ignore_index=True) if non_empty else pd.DataFrame()
    sm = am.groupby(keys, as_index=False).sum(numeric_only=False)
    tmp = sm[keys].copy()
    tx = [c for c in mp.columns if c not in keys and c not in back]
    tmp["_tot"] = mp[tx].sum(axis=1)
    tmp["_sales"] = mp["销售回款"] if "销售回款" in mp.columns else 0.0
    sm = sm.merge(tmp, on=keys, how="left")
    sm["退货率"] = sm.apply(
        lambda r: r["退款订单数量"] / r["回款订单数量"] if r["回款订单数量"] else 0.0,
        axis=1,
    )
    sm["退货金额比例"] = sm.apply(
        lambda r: r["退货总额"] / r["_sales"] if r["_sales"] else 0.0, axis=1
    )
    sm["赔付率"] = sm.apply(
        lambda r: r["赔付订单数量"] / r["回款订单数量"] if r["回款订单数量"] else 0.0,
        axis=1,
    )
    sm["赔付金额比例"] = sm.apply(
        lambda r: r["赔付金额"] / r["_tot"] if r["_tot"] else 0.0, axis=1
    )
    sm = sm.drop(columns=["_tot", "_sales"])
    return mp, sm


def seller_center_data(folder: Path):
    files = list(folder.glob("*.xlsx"))
    path = None
    for f in files:
        if "卖家中心" in f.name:
            path = f
            break
    if path is None:
        path = folder / "卖家中心账务明细.xlsx"
    fees = {
        "仓储综合服务费": 0.0,
        "EPR费用": 0.0,
        "推广服务费": 0.0,
        "发货面单费": 0.0,
        "退货面单费": 0.0,
        "提现总额": 0.0,
    }
    check = {"结算总额": 0.0, "支出总额": 0.0, "调整总额": 0.0}
    subsidy_adj = 0.0
    if not path.exists():
        print("[WARN] 未找到卖家中心账务明细.xlsx，店铺费用项将标记为0")
        return fees, check, subsidy_adj, 0.0, 0.0

    df = pd.read_excel(path, sheet_name="账务明细列表")
    df["收支金额"] = pd.to_numeric(df["收支金额"], errors="coerce").fillna(0)
    df["备注"] = df["备注"].astype(str)

    check["结算总额"] = float(df.loc[df["账务类型"] == "结算", "收支金额"].sum())
    check["支出总额"] = float(df.loc[df["账务类型"] == "支出", "收支金额"].sum())
    check["调整总额"] = float(df.loc[df["账务类型"] == "调整", "收支金额"].sum())
    fees["提现总额"] = float(df.loc[df["账务类型"] == "提现", "收支金额"].sum())

    for key, kw in [
        ("仓储综合服务费", "仓储"),
        ("推广服务费", "推广服务费"),
        ("EPR费用", "EPR"),
        ("发货面单费", "发货面单费"),
        ("退货面单费", "退货面单费"),
    ]:
        fees[key] = abs(
            float(
                df.loc[
                    (df["账务类型"] == "支出") & (df["备注"].str.contains(kw)),
                    "收支金额",
                ].sum()
            )
        )

    subsidy_adj = abs(
        float(
            df.loc[
                (df["账务类型"] == "支出")
                & (df["备注"].str.contains("非商责平台售后补贴调整")),
                "收支金额",
            ].sum()
        )
    )
    total_expense = abs(float(df.loc[df["账务类型"] == "支出", "收支金额"].sum()))
    comp_from_center = abs(
        float(
            df.loc[
                (df["账务类型"] == "支出") & (df["备注"].str.contains("履约保障")),
                "收支金额",
            ].sum()
        )
    )
    return fees, check, subsidy_adj, total_expense, comp_from_center


def append_region_excel(p, m, out_path: Path, company=None, shop=None):
    """以追加模式写入区域汇总，避免同一店铺重复。"""
    text_cols_map = {
        "SKU回款汇总": {"公司名", "店铺名", "SKU ID", "SKU货号", "货品名称", "SKU属性"},
        "SKU指标分析": {"公司名", "店铺名", "SKU ID", "SKU货号", "货品名称", "SKU属性"},
    }

    if company and shop:
        p = p.copy()
        m = m.copy()
        p.insert(0, "店铺名", shop)
        p.insert(0, "公司名", company)
        m.insert(0, "店铺名", shop)
        m.insert(0, "公司名", company)
    sheets = {"SKU回款汇总": p, "SKU指标分析": m}

    # 追加模式
    if out_path.exists():
        xl = pd.ExcelFile(str(out_path))
        existing = {s: pd.read_excel(xl, sheet_name=s) for s in xl.sheet_names}

        # 检查重复
        if "SKU回款汇总" in existing:
            df_check = existing["SKU回款汇总"]
            if "公司名" in df_check.columns and "店铺名" in df_check.columns:
                mask = (df_check["公司名"] == company) & (df_check["店铺名"] == shop)
                if mask.any():
                    return False  # 已存在，跳过

        for name, df in sheets.items():
            if name in existing:
                old = existing[name]
                # 统一列结构：缺失列填充 0，并按新数据列序输出
                for col in df.columns:
                    if col not in old.columns:
                        old[col] = 0
                for col in old.columns:
                    if col not in df.columns:
                        df[col] = 0
                df = df[old.columns]
                to_concat = [x for x in [old, df] if not x.empty]
                existing[name] = pd.concat(to_concat, ignore_index=True) if to_concat else pd.DataFrame()
            else:
                existing[name] = df
    else:
        existing = sheets

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for name, df in existing.items():
            if df.empty:
                continue
            df.to_excel(writer, index=False, sheet_name=name)
            numeric_cols = [
                c for c in df.columns if c not in text_cols_map.get(name, set())
            ]
            fmt_worksheet(writer.sheets[name], numeric_cols, df)
    return True


def append_summary(out_path: Path, company, shop, df_summary, df_expense, df_sku):
    sheets = {
        "店铺汇总数据": df_summary,
        "店铺支出统计": df_expense,
        "sku数据": df_sku,
    }
    text_cols_map = {
        "店铺汇总数据": {"公司名称", "店铺名称", "区域"},
        "店铺支出统计": {"公司名", "店铺"},
        "sku数据": {"企业名", "店铺名", "SKU ID", "SKU货号", "货品名称", "SKU属性"},
    }

    if out_path.exists():
        xl = pd.ExcelFile(str(out_path))
        existing = {s: pd.read_excel(xl, sheet_name=s) for s in xl.sheet_names}

        if "店铺汇总数据" in existing:
            df_check = existing["店铺汇总数据"]
            if "公司名称" in df_check.columns and "店铺名称" in df_check.columns:
                mask = (df_check["公司名称"] == company) & (df_check["店铺名称"] == shop)
                if mask.any():
                    return False

        for name, df in sheets.items():
            if name in existing:
                old = existing[name]
                # 统一列结构：缺失列填充 0，并按新数据列序输出
                for col in df.columns:
                    if col not in old.columns:
                        old[col] = 0
                for col in old.columns:
                    if col not in df.columns:
                        df[col] = 0
                df = df[old.columns]
                to_concat = [x for x in [old, df] if not x.empty]
                existing[name] = pd.concat(to_concat, ignore_index=True) if to_concat else pd.DataFrame()
            else:
                existing[name] = df
    else:
        existing = sheets

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for name, df in existing.items():
            if df.empty:
                continue
            df.to_excel(writer, index=False, sheet_name=name)
            numeric_cols = [
                c for c in df.columns if c not in text_cols_map.get(name, set())
            ]
            fmt_worksheet(writer.sheets[name], numeric_cols, df)
            if name == "店铺汇总数据":
                ws = writer.sheets[name]
                for col_idx, col_name in enumerate(df.columns, start=1):
                    if col_name == "最终回款总额":
                        ws.cell(row=1, column=col_idx).comment = Comment(
                            "最终回款额 = 销售总额 - 退货总额 + 调整额 - 支出总额",
                            "系统",
                        )
                        break
            if name == "sku数据":
                ws = writer.sheets[name]
                for col_idx, col_name in enumerate(df.columns, start=1):
                    if col_name == "退货率":
                        for row_idx in range(2, len(df) + 2):
                            cell = ws.cell(row=row_idx, column=col_idx)
                            cell.number_format = "0.00%"
                        break
    return True


def build_region_data(pivots, metrics, extras):
    data = {}
    for r in REGIONS:
        p, m, e = pivots[r], metrics[r], extras[r]
        sub = float(p["非商责补贴"].sum()) if "非商责补贴" in p.columns else 0.0
        data[r] = {
            "销售回款总额": float(m["销售回款总额"].sum()),
            "运费回款": e.get("运费回款", 0.0),
            "非商责补贴": sub,
            "订单数量": int(m["回款订单数量"].sum()),
            "销售数量": int(m["销售数量"].sum()),
            "退货订单数": int(m["退款订单数量"].sum()),
            "退货总额": float(m["退货总额"].sum()) + abs(e.get("销售冲回_no_sku", 0.0)),
            "赔付总额": float(m["赔付金额"].sum()) + e.get("赔付总额", 0.0),
        }
    return data


def build_shop_data(region_data, fees, check, subsidy_adj, sc_total, sc_comp):
    shop = {k: sum(v[k] for v in region_data.values()) for k in region_data[REGIONS[0]]}
    shop["非商责补贴调整"] = subsidy_adj
    shop["调整额"] = check["调整总额"]
    shop.update(fees)
    shop["退货率"] = shop["退货订单数"] / shop["订单数量"] if shop["订单数量"] else 0.0
    shop["物流费用"] = shop["发货面单费"] + shop["退货面单费"]
    shop["净销售额"] = (
        shop["销售回款总额"]
        + shop["运费回款"]
        + shop["非商责补贴"]
        - shop["退货总额"]
        + shop["调整额"]
    )
    # 使用卖家中心总支出确保所有费用都被捕获，避免遗漏新费用项
    # 扣除卖家中心中已包含的“消费者及履约保障”（与各区赔付总额重复）
    shop["总支出额"] = shop["赔付总额"] + sc_total - sc_comp
    shop["平台其他费用"] = (
        shop["总支出额"]
        - shop["仓储综合服务费"]
        - shop["推广服务费"]
        - shop["物流费用"]
    )
    shop["净回款额"] = shop["净销售额"] - shop["总支出额"]
    return shop


def enrich_region_calcs(region_data):
    zeros = {
        k: 0.0
        for k in [
            "仓储综合服务费",
            "EPR费用",
            "推广服务费",
            "物流费用",
            "提现总额",
            "非商责补贴调整",
            "调整额",
        ]
    }
    for d in region_data.values():
        d.update(zeros)
        d["净销售额"] = (
            d["销售回款总额"]
            + d["运费回款"]
            + d["非商责补贴"]
            - d["退货总额"]
            + d["调整额"]
        )
        d["总支出额"] = (
            d["赔付总额"]
            + d["仓储综合服务费"]
            + d["EPR费用"]
            + d["推广服务费"]
            + d["物流费用"]
            + d["非商责补贴调整"]
        )
        d["平台其他费用"] = (
            d["总支出额"] - d["仓储综合服务费"] - d["推广服务费"] - d["物流费用"]
        )
        d["净回款额"] = d["净销售额"] - d["总支出额"]
    return region_data


def reconcile(shop, check):
    expected = check["结算总额"] + check["支出总额"] + check["调整总额"]
    diff = abs(shop["净回款额"] - expected)
    print(f"[CHECK] 店铺净回款额: {shop['净回款额']:.2f}")
    print(f"[CHECK] 卖家中心结算+支出+调整: {expected:.2f}")
    print(f"[CHECK] 差异: {diff:.2f}")
    if diff > 0.01:
        print("[WARN] 净回款额与卖家中心结算+支出+调整存在差异，请核对数据!")
    return diff


def build_order_income_long(region_data, shop_data, company, shop):
    names = [
        "订单数量",
        "销售数量",
        "退货订单数",
        "退货率",
        "提现总额",
        "总销售额",
        "运费回款",
        "非商责补贴",
        "非商责补贴调整",
        "调整额",
        "退货总额",
        "总支出额",
        "仓储综合服务费",
        "物流费用",
        "推广服务费",
        "平台其他费用",
        "净销售额",
        "净回款额",
    ]
    key_map = {"总销售额": "销售回款总额"}

    def val(d, k):
        if k in [
            "仓储综合服务费",
            "推广服务费",
            "总支出额",
            "平台其他费用",
            "物流费用",
        ]:
            return -d[k]
        if k in ["退货总额", "非商责补贴调整"]:
            return -d[k]
        return d.get(k, d.get(key_map.get(k, k), 0.0))

    rows = []
    row = {"公司名": company, "店铺名": shop, "区域": "店铺"}
    for n in names:
        row[n] = val(shop_data, n)
    rows.append(row)

    for r in REGIONS:
        row = {"公司名": company, "店铺名": shop, "区域": r}
        src = region_data[r]
        for n in names:
            row[n] = val(src, n)
        rows.append(row)

    return pd.DataFrame(rows)


def build_expense_summary(region_data, shop_data, company, shop_name):
    """构建店铺支出统计表"""
    columns = [
        "公司名",
        "店铺",
        "总支出额",
        "仓储费",
        "推广服务费",
        "物流费用",
        "EPR费用",
        "售后费用",
    ]
    rows = [
        {
            "公司名": company,
            "店铺": shop_name,
            "总支出额": -shop_data.get("总支出额", 0),
            "仓储费": -shop_data.get("仓储综合服务费", 0),
            "推广服务费": -shop_data.get("推广服务费", 0),
            "物流费用": -shop_data.get("物流费用", 0),
            "EPR费用": -shop_data.get("EPR费用", 0),
            "售后费用": -shop_data.get("赔付总额", 0),
        }
    ]
    return pd.DataFrame(rows, columns=columns)


def build_sku_data(df, company, shop_name):
    """构建 sku 数据表"""
    rows = []
    for (sku_id, sku_no, name, attr), g in df.groupby(
        ["SKU ID", "SKU货号", "货品名称", "SKU属性"]
    ):
        back = g["交易类型"] == "销售回款"
        refund = g["交易类型"] == "销售冲回"
        other = ~g["交易类型"].isin(["销售回款", "销售冲回"])

        qty = float(g.loc[back, "数量"].sum()) if "数量" in g.columns else 0.0
        sales_val = float(g.loc[back, "金额"].sum())
        refund_val = float(g.loc[refund, "金额"].sum())
        other_val = float(g.loc[other, "金额"].sum())

        total_income = round(sales_val + refund_val + other_val, 2)
        refund_rate = round(abs(refund_val) / sales_val, 4) if sales_val else 0.0
        rows.append(
            {
                "企业名": company,
                "店铺名": shop_name,
                "SKU ID": sku_id,
                "SKU货号": sku_no,
                "货品名称": name,
                "SKU属性": attr,
                "销售数量": int(qty),
                "销售总额": round(sales_val, 2),
                "退货金额": round(refund_val, 2),
                "退货率": refund_rate,
                "其他收入": round(other_val, 2),
                "总收入": total_income,
            }
        )
    return pd.DataFrame(rows)


def build_shop_summary(region_data, shop_data, company, shop_name):
    """构建店铺汇总数据表"""
    columns = [
        "公司名称",
        "店铺名称",
        "区域",
        "总销售额",
        "订单数量",
        "销售数量",
        "退货订单数",
        "退货总额",
        "运费回款",
        "平台调整金额",
        "总支出额",
        "提现总额",
        "最终回款总额",
    ]

    rows = []

    # 店铺汇总行
    platform_adj = (
        shop_data.get("非商责补贴", 0)
        + shop_data.get("调整额", 0)
        - shop_data.get("非商责补贴调整", 0)
    )

    rows.append(
        {
            "公司名称": company,
            "店铺名称": shop_name,
            "区域": "店铺",
            "总销售额": shop_data.get("销售回款总额", 0),
            "订单数量": int(shop_data.get("订单数量", 0)),
            "销售数量": int(shop_data.get("销售数量", 0)),
            "退货订单数": int(shop_data.get("退货订单数", 0)),
            "退货总额": -shop_data.get("退货总额", 0),
            "运费回款": shop_data.get("运费回款", 0),
            "平台调整金额": platform_adj,
            "总支出额": -shop_data.get("总支出额", 0),
            "提现总额": shop_data.get("提现总额", 0),
            "最终回款总额": shop_data.get("净回款额", 0),
        }
    )

    # 各区域行
    for region in REGIONS:
        rd = region_data[region]
        rows.append(
            {
                "公司名称": company,
                "店铺名称": shop_name,
                "区域": region,
                "总销售额": rd.get("销售回款总额", 0),
                "订单数量": int(rd.get("订单数量", 0)),
                "销售数量": int(rd.get("销售数量", 0)),
                "退货订单数": int(rd.get("退货订单数", 0)),
                "退货总额": -rd.get("退货总额", 0),
                "运费回款": rd.get("运费回款", 0),
                "平台调整金额": rd.get("非商责补贴", 0),
                "总支出额": -rd.get("总支出额", 0),
                "提现总额": rd.get("提现总额", 0),
                "最终回款总额": rd.get("净回款额", 0),
            }
        )

    return pd.DataFrame(rows, columns=columns)


def main():
    parser = argparse.ArgumentParser(description="解析 Temu 各区账务明细")
    parser.add_argument("-folder", required=True, help="数据文件夹路径")
    parser.add_argument("-company", required=True, help="公司名称")
    parser.add_argument("-shop", required=True, help="店铺名称")
    args = parser.parse_args()
    folder = Path(args.folder).expanduser().resolve()
    company = args.company.strip()
    shop_name = args.shop.strip()

    if not folder.is_dir():
        print(f"[ERROR] 指定路径不是文件夹: {folder}")
        sys.exit(1)

    missing = [r for r in REGIONS if not find_region_file(folder, r).exists()]
    if missing:
        print("[ERROR] 文件夹中缺少以下必需的表格文件:")
        for f in missing:
            print(f"    - {f}")
        sys.exit(1)

    out_dir = Path.cwd() / "temu汇总"
    out_dir.mkdir(exist_ok=True)
    out_path = out_dir / "汇总.xlsx"

    pivots, metrics, extras, settles = {}, {}, {}, {}
    for r in REGIONS:
        region_path = find_region_file(folder, r)
        print(f"[INFO] 正在处理 {r} 数据: {region_path.name}")
        p, m, e, s = process_region(region_path)
        pivots[r] = p
        metrics[r] = m
        extras[r] = e
        settles[r] = s
        appended = append_region_excel(p, m, out_dir / f"{r}汇总.xlsx", company, shop_name)
        if not appended:
            print(f"[WARN] {r}汇总.xlsx 中已存在 {company}/{shop_name} 的数据，跳过。")

    print("[INFO] 正在生成店铺汇总数据 ...")
    region_data = enrich_region_calcs(build_region_data(pivots, metrics, extras))
    fees, check, subsidy_adj, sc_total, sc_comp = seller_center_data(folder)
    shop_data = build_shop_data(region_data, fees, check, subsidy_adj, sc_total, sc_comp)
    reconcile(shop_data, check)

    # 合并所有区域结算数据
    all_settle = pd.concat([s for s in settles.values() if not s.empty], ignore_index=True)
    df_sku = build_sku_data(all_settle, company, shop_name)

    df_summary = build_shop_summary(region_data, shop_data, company, shop_name)
    df_expense = build_expense_summary(region_data, shop_data, company, shop_name)

    ok = append_summary(out_path, company, shop_name, df_summary, df_expense, df_sku)
    if not ok:
        print(f"[WARN] 汇总文件中已存在 {company}/{shop_name} 的数据，跳过写入。")
    else:
        print(f"[INFO] 已追加到汇总文件: {out_path.resolve()}")

    print(f"\n全部处理完成，输出目录: {out_dir.resolve()}")


if __name__ == "__main__":
    main()
